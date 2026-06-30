"""
Read ISBNs from Excel, CSV, or plain text input files.

Supported inputs:
- Excel: .xlsx
- CSV: .csv
- Text: .txt, .list, or any plain text file

The reader searches the entire file for ISBN-10 and ISBN-13 values.
It does not require specific column names.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from models import SourceRow


ISBN_PATTERN = re.compile(
    r"""
    (?:
        97[89][-\s]?
        \d[-\s]?
        \d{2,5}[-\s]?
        \d{2,7}[-\s]?
        \d
    )
    |
    (?:
        \d[-\s]?
        \d{2,5}[-\s]?
        \d{2,7}[-\s]?
        [\dXx]
    )
    """,
    re.VERBOSE,
)


def normalize_isbn(value: str) -> str:
    return re.sub(r"[^0-9Xx]", "", value).upper()


def looks_like_isbn(value: str) -> bool:
    isbn = normalize_isbn(value)
    return len(isbn) in (10, 13)


class ExcelReader:
    def read(self, filename: str | Path) -> list[SourceRow]:
        path = Path(filename)

        if not path.exists():
            raise FileNotFoundError(path)

        suffix = path.suffix.lower()

        if suffix == ".xlsx":
            return self._read_excel(path)

        if suffix == ".csv":
            return self._read_csv(path)

        return self._read_text(path)

    def _make_rows(self, found: list[tuple[int, str, str]]) -> list[SourceRow]:
        rows: list[SourceRow] = []

        for index, (line_number, isbn, raw_text) in enumerate(found, start=1):
            rows.append(
                SourceRow(
                    row_number=line_number,
                    copy_number=index,
                    isbn=isbn,
                    raw={"source_text": raw_text},
                )
            )

        return rows

    def _extract_from_text(self, text: str, line_number: int) -> list[tuple[int, str, str]]:
        found: list[tuple[int, str, str]] = []

        for match in ISBN_PATTERN.findall(text):
            isbn = normalize_isbn(match)
            if looks_like_isbn(isbn):
                found.append((line_number, isbn, text.strip()))

        return found

    def _read_text(self, path: Path) -> list[SourceRow]:
        found: list[tuple[int, str, str]] = []

        with path.open("r", encoding="utf-8-sig") as file:
            for line_number, line in enumerate(file, start=1):
                found.extend(self._extract_from_text(line, line_number))

        return self._make_rows(found)

    def _read_csv(self, path: Path) -> list[SourceRow]:
        found: list[tuple[int, str, str]] = []

        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.reader(file)

            for row_number, row in enumerate(reader, start=1):
                row_text = ",".join(str(cell) for cell in row if cell is not None)
                found.extend(self._extract_from_text(row_text, row_number))

        return self._make_rows(found)

    def _read_excel(self, path: Path) -> list[SourceRow]:
        found: list[tuple[int, str, str]] = []

        workbook = load_workbook(path, read_only=True, data_only=True)

        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                found.extend(self._extract_from_text(row_text, row_number))

        workbook.close()

        return self._make_rows(found)
